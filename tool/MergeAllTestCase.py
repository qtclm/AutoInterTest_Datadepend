import openpyxl
import os

class OperationDataCase(object):
    def __init__(self, file_address=None, sheet_id=None,path='../dataCase'):
        newfile_path=path+'/run'
        if not os.path.exists(newfile_path):
            print('文件夹不存在，开始创建文件夹:"%s"' % (newfile_path))
            os.makedirs(newfile_path)
            print('文件夹"%s"已创建' % (newfile_path))
        self.new_file=os.path.join(newfile_path,'DataCase_ALL.xlsx')
        if not os.path.exists(self.new_file):
            print('文件不存在，开始创建文件:"%s"' % (self.new_file))
            with open(self.new_file, 'w') as file:  # 创建文件
                file.close()
            print('文件夹"%s"已创建' % (self.new_file))
        self.dataCase_list = []
        if file_address is not None:
            self.file_address=file_address
            self.sheet_id=sheet_id
            self.dataCase_list.append(self.file_address)
        else:
            for test_list in os.listdir(path):
                if test_list[-5:]=='.xlsx':
                    file_address = path + '/' + test_list
                    self.file_address = file_address
                    self.sheet_id = sheet_id
                    sheet_id = 0
                    self.dataCase_list.append(self.file_address)
        self.data = self.get_dataBatch()
        self.caseidrow_List,self.dependCaseidrow_List=[],[]#存放caseid最终得行数与dependcaseid最终得行号
        self.write_data_to_excle()


    # 获取单元格的行数
    def get_lines(self):
        tables = self.data
        lines_list=[]
        if tables :
            for lines in tables:
                '''max_row:获取行数，max_column:获取列数'''
                lines_list.append(lines.max_row)
            return lines_list
        else:
            print("没有找到excle文件")
            return None

    #获取目录下所有excle的页签
    def get_dataBatch(self):
        tables_list = []
        dataCase_list=self.dataCase_list
        if  dataCase_list:
            for dataCase in dataCase_list:
                # read_only=True:设置工作薄属性为只读
                data=openpyxl.load_workbook(dataCase)
                tables=data.sheetnames
                for table in tables:
                    table_data=data[table]
                    tables_list.append(table_data)
            return tables_list
        else:
            print("没有找到excle文件")
            return None

    # 根据行号找到该行内容
    def get_row_values(self):
        tables = self.data
        sheet_rows = self.get_lines()
        row_datas=[]
        row_count=0
        if  tables :
            for table in tables:
                for row in range(1,max(sheet_rows)+1):
                    row_data=table[row]
                    row_list=[i.value for i in row_data]
                    if (row_list[0]==None) or (row==1 and (row_list in row_datas)) :
                        #第一个单元格内容为None，说明改行没有数据，直接过滤
                        # 如果当前行是1行且数据已存在与exlce中，过滤，防止表头写入重复
                        continue
                    else:
                        row_count += 1
                        row_datas.append(row_list)
                        if row_list[7] and row!=1:
                            # dependcaseid对应得行号：caseId最终得index-（最初得caseId-最初得dependCaseId)，还需在减一（因为caseId是从第二行才开始）
                            dependCaseidrow=row_count-(row_list[0]-row_list[7])-1
                            self.dependCaseidrow_List.append(dependCaseidrow)
                            self.caseidrow_List.append(row_count)
                        
            
            # print(row_datas)
            # print(self.caseidrow_List)
            # print(self.dependCaseidrow_List)
            return row_datas

        else:
            print("当前路径没有可运行的excle文件")
            return None

    def write_data_to_excle(self):
        datas=self.get_row_values()
        # print(datas)
        num=max([len(i) for i in datas])#获取最大列
        # write_only = True：设置工作薄属性为只写
        new_book=openpyxl.Workbook()
        sheet=new_book.active
        sheet.title = "datacase"
        for index,data in enumerate(datas):
            # 将列表的数据逐行写入excle
            sheet.append(data)
            # '''处理CaseId'''
            if index>1:
                sheet['A{}'.format(index+1)].value=index
            # 处理dependCaseId
            if index+1 in self.caseidrow_List:
                caseId_index=self.caseidrow_List.index(index+1)
                # column对应dependCaseId对应的列
                sheet.cell(row=index+1,column=8).value=self.dependCaseidrow_List[caseId_index]
                # print(sheet.cell(row=index+1,column=7).value)
        new_book.save(self.new_file)



if __name__=="__main__":
    op=OperationDataCase()
    # op.get_data()
    # op.get_dataBatch()
    # op.get_row_values()
    # print(a)
    # op.write_data_to_excle()